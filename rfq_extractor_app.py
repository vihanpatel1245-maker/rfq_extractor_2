import streamlit as st
import pdfplumber
import re
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import json

# ── Page config (must be first st command) ────────────────────────────────────
st.set_page_config(page_title="RFQ PDF → Excel", page_icon="📋", layout="centered")

# ── Authentication ─────────────────────────────────────────────────────────────
USERS = {"anany": "dada.niruma"}

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False


def login_screen():
    # ── Glassmorphism Login CSS ───────────────────────────────────────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    /* Hide sidebar, header, footer, deploy button on login */
    [data-testid="stSidebar"],
    [data-testid="stToolbar"],
    footer,
    #MainMenu { display: none !important; }

    /* Hide "Press Enter to submit form" helper text */
    [data-testid="InputInstructions"],
    .stFormSubmitButton [data-testid="InputInstructions"] {
        display: none !important;
    }

    header[data-testid="stHeader"] {
        background: transparent !important;
    }

    /* Animated gradient background */
    .stApp {
        background: linear-gradient(
            135deg,
            #1a0533 0%,
            #2d1b69 20%,
            #0f2027 40%,
            #3a1c71 60%,
            #1a0533 80%,
            #d76d77 100%
        );
        background-size: 400% 400%;
        animation: gradientShift 12s ease infinite;
        font-family: 'Inter', sans-serif;
    }

    @keyframes gradientShift {
        0%   { background-position: 0% 50%; }
        50%  { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }

    /* Center content vertically */
    [data-testid="stMain"] {
        display: flex;
        align-items: center;
        justify-content: center;
        min-height: 100vh;
    }

    [data-testid="stMainBlockContainer"] {
        max-width: 420px !important;
        width: 100%;
        padding-top: 0 !important;
    }

    /* Glass card for the form */
    [data-testid="stForm"] {
        background: rgba(255, 255, 255, 0.07) !important;
        backdrop-filter: blur(24px) saturate(180%) !important;
        -webkit-backdrop-filter: blur(24px) saturate(180%) !important;
        border: 1px solid rgba(255, 255, 255, 0.12) !important;
        border-radius: 24px !important;
        padding: 1.5rem 2rem 2rem !important;
        box-shadow:
            0 8px 32px rgba(0, 0, 0, 0.35),
            inset 0 1px 0 rgba(255, 255, 255, 0.1) !important;
    }

    /* Input fields — underline style */
    [data-testid="stForm"] [data-testid="stTextInput"] > div > div {
        background: rgba(255, 255, 255, 0.04) !important;
        border: none !important;
        border-bottom: 1.5px solid rgba(255, 255, 255, 0.25) !important;
        border-radius: 0 !important;
        transition: border-color 0.3s ease !important;
    }

    [data-testid="stForm"] [data-testid="stTextInput"] > div > div:focus-within {
        border-bottom-color: rgba(215, 109, 119, 0.8) !important;
        box-shadow: none !important;
    }

    [data-testid="stForm"] input {
        color: rgba(255, 255, 255, 0.9) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.95rem !important;
        padding: 0.7rem 0.5rem !important;
        caret-color: #d76d77 !important;
    }

    [data-testid="stForm"] input::placeholder {
        color: rgba(255, 255, 255, 0.35) !important;
    }

    /* Labels */
    [data-testid="stForm"] label {
        color: rgba(255, 255, 255, 0.7) !important;
        font-weight: 400 !important;
        font-size: 0.85rem !important;
        letter-spacing: 0.5px !important;
    }

    /* LOGIN button */
    [data-testid="stForm"] button[kind="primary"] {
        background: linear-gradient(
            135deg,
            #1a0533 0%,
            #2d1b69 40%,
            #3a1c71 70%,
            #0f2027 100%
        ) !important;
        border: 1px solid rgba(255, 255, 255, 0.15) !important;
        border-radius: 30px !important;
        color: white !important;
        padding: 0.8rem 2rem !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        letter-spacing: 3px !important;
        text-transform: uppercase !important;
        transition: all 0.35s ease !important;
        margin-top: 0.8rem !important;
        box-shadow: 0 4px 15px rgba(26, 5, 51, 0.4) !important;
    }

    [data-testid="stForm"] button[kind="primary"]:hover {
        background: linear-gradient(
            135deg,
            #2d1b69 0%,
            #4a2d8a 40%,
            #d76d77 100%
        ) !important;
        box-shadow: 0 6px 20px rgba(45, 27, 105, 0.5) !important;
        transform: translateY(-2px);
        border-color: rgba(255, 255, 255, 0.25) !important;
    }

    [data-testid="stForm"] button[kind="primary"]:active {
        transform: translateY(0);
    }

    /* Error message styling */
    [data-testid="stForm"] + div .stAlert {
        background: rgba(255, 75, 75, 0.15) !important;
        border: 1px solid rgba(255, 75, 75, 0.3) !important;
        border-radius: 12px !important;
        color: #ff8a8a !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Avatar + Title (above form, centered) ─────────────────────────────────
    st.markdown("""
    <div style="text-align: center; margin-bottom: 0.5rem; margin-top: -2rem;">
        <div style="
            width: 88px;
            height: 88px;
            border-radius: 50%;
            background: linear-gradient(135deg, rgba(255,255,255,0.12), rgba(255,255,255,0.04));
            border: 2px solid rgba(255, 255, 255, 0.15);
            margin: 0 auto 0.8rem;
            display: flex;
            align-items: center;
            justify-content: center;
            backdrop-filter: blur(10px);
            box-shadow: 0 4px 20px rgba(0,0,0,0.2);
        ">
            <svg width="40" height="40" viewBox="0 0 24 24" fill="none" style="opacity: 0.5;">
                <path d="M12 12c2.7 0 4.8-2.1 4.8-4.8S14.7 2.4 12 2.4 7.2 4.5 7.2 7.2 9.3 12 12 12zm0 2.4c-3.2 0-9.6 1.6-9.6 4.8v1.2c0 .7.5 1.2 1.2 1.2h16.8c.7 0 1.2-.5 1.2-1.2v-1.2c0-3.2-6.4-4.8-9.6-4.8z" fill="rgba(255,255,255,0.6)"/>
            </svg>
        </div>
        <p style="
            color: rgba(255, 255, 255, 0.6);
            font-family: 'Inter', sans-serif;
            font-size: 0.8rem;
            letter-spacing: 2px;
            text-transform: uppercase;
            margin: 0;
        ">Welcome Back</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Login Form ────────────────────────────────────────────────────────────
    with st.form("login_form"):
        username = st.text_input("✉  Username", placeholder="Enter your username", autocomplete="off")
        password = st.text_input("🔒  Password", type="password", placeholder="Enter your password", autocomplete="new-password")
        submitted = st.form_submit_button("LOGIN", type="primary", use_container_width=True)

    # Disable browser autofill/password suggestions via JS
    import streamlit.components.v1 as components
    components.html("""
    <script>
    (function() {
        const doc = window.parent.document;
        // Disable autocomplete on all inputs and the form itself
        const forms = doc.querySelectorAll('form');
        forms.forEach(f => {
            f.setAttribute('autocomplete', 'off');
        });
        const inputs = doc.querySelectorAll('input');
        inputs.forEach(input => {
            input.setAttribute('autocomplete', 'off');
            input.setAttribute('data-form-type', 'other');
            input.setAttribute('data-lpignore', 'true');
            input.setAttribute('data-1p-ignore', 'true');
            // Randomize name to confuse browser password manager
            input.setAttribute('name', 'field_' + Math.random().toString(36).substr(2, 9));
        });
    })();
    </script>
    """, height=0)

    if submitted:
        if username in USERS and USERS[username] == password:
            st.session_state.authenticated = True
            st.session_state.username = username
            st.rerun()
        else:
            st.error("❌ Invalid username or password")


if not st.session_state.authenticated:
    login_screen()
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP (after login)
# ══════════════════════════════════════════════════════════════════════════════

# ── Main app styles ───────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* Hide sidebar completely */
[data-testid="stSidebar"] { display: none !important; }

/* Same animated gradient background as login */
.stApp {
    background: linear-gradient(
        135deg,
        #1a0533 0%,
        #2d1b69 20%,
        #0f2027 40%,
        #3a1c71 60%,
        #1a0533 80%,
        #d76d77 100%
    );
    background-size: 400% 400%;
    animation: gradientShift 12s ease infinite;
    font-family: 'Inter', sans-serif;
}

@keyframes gradientShift {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}

/* Header transparent */
header[data-testid="stHeader"] {
    background: transparent !important;
}

/* Page title */
h1 {
    font-family: 'Inter', sans-serif !important;
    font-weight: 700 !important;
    color: white !important;
    letter-spacing: -0.5px !important;
}

/* Subheaders */
h2, h3 {
    font-family: 'Inter', sans-serif !important;
    color: rgba(255, 255, 255, 0.9) !important;
}

/* Body text */
p, span, label, .stCaption, [data-testid="stText"] {
    font-family: 'Inter', sans-serif !important;
    color: rgba(255, 255, 255, 0.8) !important;
}

/* File uploader — glass card */
[data-testid="stFileUploader"] {
    background: rgba(255, 255, 255, 0.06) !important;
    backdrop-filter: blur(20px) saturate(160%) !important;
    -webkit-backdrop-filter: blur(20px) saturate(160%) !important;
    border: 1px solid rgba(255, 255, 255, 0.1) !important;
    border-radius: 16px !important;
    padding: 1.2rem !important;
    box-shadow: 0 4px 20px rgba(0, 0, 0, 0.2) !important;
}

[data-testid="stFileUploader"] label {
    color: rgba(255, 255, 255, 0.7) !important;
    font-weight: 500 !important;
    letter-spacing: 0.3px !important;
}

/* Upload button inside uploader */
[data-testid="stFileUploader"] button {
    background: rgba(255, 255, 255, 0.1) !important;
    border: 1px solid rgba(255, 255, 255, 0.15) !important;
    color: rgba(255, 255, 255, 0.8) !important;
    border-radius: 8px !important;
    backdrop-filter: blur(10px) !important;
}

[data-testid="stFileUploader"] button:hover {
    background: rgba(255, 255, 255, 0.18) !important;
    border-color: rgba(255, 255, 255, 0.25) !important;
}

[data-testid="stFileUploader"] small {
    color: rgba(255, 255, 255, 0.4) !important;
}



/* Uploaded file chips */
[data-testid="stFileUploader"] [data-testid="stUploadedFile"] {
    background: rgba(255, 255, 255, 0.08) !important;
    border: 1px solid rgba(255, 255, 255, 0.12) !important;
    border-radius: 10px !important;
}

/* Dataframe / preview — glass card */
[data-testid="stDataFrame"] {
    background: rgba(255, 255, 255, 0.05) !important;
    backdrop-filter: blur(16px) !important;
    border: 1px solid rgba(255, 255, 255, 0.08) !important;
    border-radius: 14px !important;
    overflow: hidden !important;
    box-shadow: 0 4px 16px rgba(0, 0, 0, 0.15) !important;
}

/* Primary buttons */
div.stButton > button[kind="primary"],
div.stDownloadButton > button[kind="primary"] {
    background: linear-gradient(135deg, #0E577C 0%, #1a7aad 100%) !important;
    border: 1px solid rgba(255, 255, 255, 0.12) !important;
    color: white !important;
    border-radius: 12px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    letter-spacing: 0.5px !important;
    padding: 0.6rem 1.5rem !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 4px 15px rgba(14, 87, 124, 0.3) !important;
}
div.stButton > button[kind="primary"]:hover,
div.stDownloadButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #0a3f5a 0%, #0E577C 100%) !important;
    border-color: rgba(255, 255, 255, 0.2) !important;
    box-shadow: 0 6px 20px rgba(14, 87, 124, 0.4) !important;
    transform: translateY(-1px);
}

/* Progress bar */
[data-testid="stProgress"] > div > div {
    background-color: rgba(255, 255, 255, 0.1) !important;
    border-radius: 10px !important;
}
[data-testid="stProgress"] > div > div > div {
    background: linear-gradient(90deg, #0E577C, #1a7aad, #d76d77) !important;
    border-radius: 10px !important;
}

/* Divider */
hr {
    border-color: rgba(255, 255, 255, 0.1) !important;
}

/* Spinner text */
.stSpinner > div {
    color: rgba(255, 255, 255, 0.7) !important;
}

/* Caption */
[data-testid="stCaption"] {
    color: rgba(255, 255, 255, 0.5) !important;
}
</style>
""", unsafe_allow_html=True)

# ── Page Title ────────────────────────────────────────────────────────────────
st.title("PDF → Excel")

# ── Columns ───────────────────────────────────────────────────────────────────
COLUMNS = [
    "Enquiry", "Cust", "RFQ D", "Due D", "Due Time",
    "Delivery Time", "Status", "SN", "CPN", "Description",
    "Enq. Part No*", "Enq. Mfg", "Qty", "Unit",
    "ITEM/VALUE.(EVALUATION)", "TP", "Remark", "ACT.Due DT"
]

EXCEL_DATE_BASE = datetime(1899, 12, 30)

def to_serial(d: datetime) -> int:
    return (d - EXCEL_DATE_BASE).days

# ── ITEM/VALUE.(EVALUATION) extractor ────────────────────────────────────────
def get_item_value(text: str, cust_code: str) -> str:
    code = cust_code.upper()
    if "N6" in code:
        m = re.search(
            r"(Year of manufacturing[^.]*?within\s+\d+\s+years?[^.]*\.)",
            text, re.IGNORECASE
        )
        if m:
            return m.group(1).strip()
    if "CH1" in code:
        m = re.search(
            r"(MANUFACTURING DATE CODE[^.\n]*?WITHIN\s+\d+\s+YEARS?[^.\n]*(?:IS MANDATORY|ORDER)[^.\n]*)",
            text, re.IGNORECASE
        )
        if m:
            return m.group(1).strip()
    return "N/A"

# ── Core extraction ───────────────────────────────────────────────────────────
def extract_from_pdf(pdf_bytes: bytes) -> list[dict]:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    # 1. Enquiry
    enq_m = re.search(r"RFx number\s+(\d+)", text)
    enquiry = enq_m.group(1) if enq_m else "NA"

    # 2. Cust
    desc_m = re.search(r"Description:\s*(\S+)", text)
    cust = "BEL"
    if desc_m:
        parts = desc_m.group(1).split("/")
        cust = "BEL-" + parts[1] if len(parts) > 1 else "BEL-" + parts[0]

    # 2b. ITEM/VALUE
    item_value = get_item_value(text, cust)

    # 3. RFQ D
    rfq_d = datetime.today()

    # 4 & 5 & 18. Submission period
    sub_line = re.search(r"Submission period:\s*(.+)", text, re.IGNORECASE)
    if sub_line:
        all_dates = re.findall(r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2})", sub_line.group(1))
        if all_dates:
            last_date, last_time = all_dates[-1]
            sub_date = datetime.strptime(last_date, "%d.%m.%Y")
            due_d    = sub_date - timedelta(days=3)
            t        = datetime.strptime(last_time, "%H:%M")
            due_time = t.strftime("%I:%M %p").lstrip("0")
            act_due  = sub_date
        else:
            due_d = due_time = act_due = None
    else:
        due_d = due_time = act_due = None

    # 6. Delivery Time
    lt_m = re.search(
        r"(?:OUR\s+REQUIRED\s+)?DELIVERY\s+SCHDULE?\s*:?\s*([\d\-–]+\s*(?:WEEKS?|DAYS?))",
        text, re.IGNORECASE
    )
    delivery_time = "NA"
    if lt_m:
        lt_raw  = lt_m.group(1).strip()
        range_m = re.match(r"(\d+)[\-–](\d+)\s*DAYS?", lt_raw, re.IGNORECASE)
        days_m  = re.match(r"(\d+)\s*DAYS?",            lt_raw, re.IGNORECASE)
        weeks_m = re.match(r"([\d\-–]+)\s*WEEKS?",      lt_raw, re.IGNORECASE)
        if range_m:
            avg = (int(range_m.group(1)) + int(range_m.group(2))) / 2
            delivery_time = f"{round(avg / 7)} WEEKS"
        elif days_m:
            delivery_time = f"{round(int(days_m.group(1)) / 7)} WEEKS"
        elif weeks_m:
            delivery_time = weeks_m.group(1) + " WEEKS"

    # 7+. Bid Details
    NOISE = re.compile(
        r"^(?:Item\s+Material|Qty/Unit|Quantity\s*$|Page\s+\d|Date\s*:|"
        r"Bid Invitation|Product no\.|info@|nklamin@|orantselectro@|"
        r"sales\.|support\.|\d{7,10}$)",
        re.IGNORECASE
    )
    in_bid    = False
    bid_lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if re.search(r"^Bid Details$", s):
            in_bid = True
            continue
        if in_bid:
            if NOISE.match(s):
                continue
            bid_lines.append(s)

    rows = []
    i = 0
    while i < len(bid_lines):
        item_m = re.match(
            r"^(\d+)\s+(\d{8,})\s+(.+?)\s+([\d,\.]+)\s+([A-Z]+)$",
            bid_lines[i]
        )
        if item_m:
            sn      = int(item_m.group(1)) // 10
            cpn     = item_m.group(2)
            desc    = item_m.group(3).strip()
            qty_raw = item_m.group(4).replace(",", "")
            try:
                qty = float(qty_raw)
                qty = int(qty) if qty == int(qty) else qty
            except:
                qty = qty_raw
            unit = item_m.group(5)

            mfg_list, mpn_list = [], []
            while i + 1 < len(bid_lines):
                nxt = bid_lines[i + 1]
                if re.match(r"^\d+\s+\d{8,}", nxt):
                    break
                if "-" not in nxt and "/" not in nxt:
                    break
                nxt_n   = re.sub(r",\s*-", "-", nxt)
                split_m = re.match(r"^(.*?[A-Z\)\/\.])\s*-\s*(.+)$", nxt_n)
                if split_m:
                    mfg_list.append(split_m.group(1).strip().rstrip(",/").strip())
                    mpn_list.append(split_m.group(2).strip())
                else:
                    break
                i += 1

            mfg = " // ".join(mfg_list) if mfg_list else "NA"
            mpn = " // ".join(mpn_list) if mpn_list else "NA"

            rows.append({
                "Enquiry":                 enquiry,
                "Cust":                    cust,
                "RFQ D":                   rfq_d,
                "Due D":                   due_d,
                "Due Time":                due_time or "NA",
                "Delivery Time":           delivery_time,
                "Status":                  "Working",
                "SN":                      sn,
                "CPN":                     cpn,
                "Description":             desc,
                "Enq. Part No*":           mpn,
                "Enq. Mfg":                mfg,
                "Qty":                     qty,
                "Unit":                    unit,
                "ITEM/VALUE.(EVALUATION)": item_value,
                "TP":                      "",
                "Remark":                  "N/A",
                "ACT.Due DT":              act_due,
            })
        i += 1

    if not rows:
        rows.append({col: "NA" for col in COLUMNS} | {
            "Enquiry": enquiry, "Cust": cust,
            "RFQ D": rfq_d, "Due D": due_d,
            "Due Time": due_time or "NA",
            "Delivery Time": delivery_time,
            "Status": "Working",
            "ITEM/VALUE.(EVALUATION)": item_value,
            "Remark": "N/A", "ACT.Due DT": act_due, "TP": "",
        })

    return rows


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(all_rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFQ Log"

    hdr_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="BBBBBB")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_w    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    fill_b    = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    data_font = Font(name="Arial", size=10)
    d_align   = Alignment(horizontal="left", vertical="center")

    col_widths = [14, 10, 12, 12, 10, 14, 10, 5, 16, 40, 16, 28, 6, 6, 22, 8, 10, 12]
    date_cols  = {"RFQ D", "Due D", "ACT.Due DT"}

    for ci, (col, w) in enumerate(zip(COLUMNS, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(all_rows, 2):
        fill = fill_w if ri % 2 == 0 else fill_b
        for ci, col in enumerate(COLUMNS, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci)
            if col in date_cols:
                if isinstance(val, datetime):
                    cell.value = to_serial(val)
                    cell.number_format = "DD-MMM-YY"
                else:
                    cell.value = str(val) if val else ""
            elif col == "SN":
                try:    cell.value = int(val)
                except: cell.value = val
            elif col == "Qty":
                try:
                    f = float(val)
                    cell.value = int(f) if f == int(f) else f
                except: cell.value = val
            else:
                cell.value = str(val) if val not in (None, "") else ""
            cell.font = data_font; cell.fill = fill
            cell.alignment = d_align; cell.border = bdr
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Session state ─────────────────────────────────────────────────────────────
if "result_excel" not in st.session_state:
    st.session_state.result_excel = None
if "result_rows" not in st.session_state:
    st.session_state.result_rows  = []
if "result_count" not in st.session_state:
    st.session_state.result_count = 0
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

# ── UI ────────────────────────────────────────────────────────────────────────
uploaded_files = st.file_uploader(
    "Upload PDF",
    type="pdf",
    accept_multiple_files=True,
    key=f"pdf_uploader_{st.session_state.uploader_key}",
)

if uploaded_files:
    st.write(f"**{len(uploaded_files)} file(s) selected**")

    if not st.session_state.result_excel:
        # Show Generate button when no results yet
        if st.button("Generate Excel", type="primary", use_container_width=True):
            st.session_state.result_excel = None
            st.session_state.result_rows  = []
            st.session_state.result_count = 0

            all_rows = []
            prog = st.progress(0)

            for i, uf in enumerate(uploaded_files):
                with st.spinner(f"Reading {uf.name}…"):
                    try:
                        rows = extract_from_pdf(uf.read())
                        all_rows.extend(rows)
                    except Exception as e:
                        st.error(f"❌  {uf.name}: {e}")
                prog.progress((i + 1) / len(uploaded_files))

            if all_rows:
                st.session_state.result_excel = build_excel(all_rows)
                st.session_state.result_rows  = all_rows
                st.session_state.result_count = len(uploaded_files)
                st.rerun()
    else:
        # Show New Conversion button when results are ready
        if st.button("🔄 New Conversion", type="primary", use_container_width=True, key="new_conversion"):
            st.session_state.result_excel = None
            st.session_state.result_rows = []
            st.session_state.result_count = 0
            st.session_state.uploader_key += 1
            st.rerun()

# ── Download + preview ────────────────────────────────────────────────────────
if st.session_state.result_excel:
    st.markdown("---")
    st.download_button(
        label="Download RFQ_Log.xlsx",
        data=st.session_state.result_excel,
        file_name="RFQ_Log.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
        key="dl_combined",
    )

    st.subheader("Preview")
    preview_cols = ["Enquiry", "Cust", "Due D", "Due Time", "Delivery Time",
                    "SN", "CPN", "Description", "Enq. Part No*", "Enq. Mfg", "Qty", "Unit"]
    df = pd.DataFrame(st.session_state.result_rows)
    for dc in ["RFQ D", "Due D", "ACT.Due DT"]:
        if dc in df.columns:
            df[dc] = df[dc].apply(
                lambda x: x.strftime("%d-%b-%y") if isinstance(x, datetime) else str(x)
            )
    st.dataframe(
        df[[c for c in preview_cols if c in df.columns]],
        use_container_width=True
    )
    st.caption(
        f"Total: {len(st.session_state.result_rows)} row(s) "
        f"from {st.session_state.result_count} PDF(s)"
    )

# ── Logout Button (bottom-right, fixed via JS) ───────────────────────────────
if st.button("🚪 Logout", key="logout_bottom"):
    st.session_state.authenticated = False
    st.session_state.pop("username", None)
    st.session_state.result_excel = None
    st.session_state.result_rows = []
    st.session_state.result_count = 0
    st.rerun()

import streamlit.components.v1 as components
components.html("""
<script>
(function() {
    const doc = window.parent.document;
    const buttons = doc.querySelectorAll('button');
    buttons.forEach(btn => {
        const text = btn.textContent.trim();
        // Style New Conversion button — theme-matching warm coral
        if (text.includes('New Conversion')) {
            btn.style.setProperty('background', 'linear-gradient(135deg, #c74b8f 0%, #d76d77 100%)', 'important');
            btn.style.setProperty('border-color', 'rgba(255, 255, 255, 0.15)', 'important');
            btn.style.setProperty('color', 'white', 'important');
            btn.style.setProperty('box-shadow', '0 4px 15px rgba(199, 75, 143, 0.3)', 'important');
            btn.style.setProperty('transition', 'all 0.3s ease');
            btn.addEventListener('mouseenter', () => {
                btn.style.setProperty('background', 'linear-gradient(135deg, #a83d78 0%, #c75a68 100%)', 'important');
                btn.style.setProperty('box-shadow', '0 6px 20px rgba(199, 75, 143, 0.4)', 'important');
            });
            btn.addEventListener('mouseleave', () => {
                btn.style.setProperty('background', 'linear-gradient(135deg, #c74b8f 0%, #d76d77 100%)', 'important');
                btn.style.setProperty('box-shadow', '0 4px 15px rgba(199, 75, 143, 0.3)', 'important');
            });
        }
        if (text.includes('Logout')) {
            const container = btn.closest('[data-testid="stButton"]') || btn.closest('.stButton') || btn.parentElement;
            container.style.position = 'fixed';
            container.style.bottom = '24px';
            container.style.right = '28px';
            container.style.zIndex = '9999';
            container.style.width = 'auto';
            btn.style.background = 'rgba(255, 255, 255, 0.08)';
            btn.style.border = '1px solid rgba(255, 255, 255, 0.15)';
            btn.style.borderRadius = '10px';
            btn.style.color = 'rgba(255, 255, 255, 0.8)';
            btn.style.fontSize = '0.82rem';
            btn.style.padding = '0.45rem 1.1rem';
            btn.style.backdropFilter = 'blur(10px)';
            btn.style.boxShadow = '0 2px 12px rgba(0, 0, 0, 0.3)';
            btn.style.transition = 'all 0.3s ease';
            btn.style.cursor = 'pointer';
            btn.style.width = 'auto';
            btn.addEventListener('mouseenter', () => {
                btn.style.background = 'rgba(255, 75, 75, 0.2)';
                btn.style.borderColor = 'rgba(255, 75, 75, 0.4)';
                btn.style.color = '#ff8a8a';
            });
            btn.addEventListener('mouseleave', () => {
                btn.style.background = 'rgba(255, 255, 255, 0.08)';
                btn.style.borderColor = 'rgba(255, 255, 255, 0.15)';
                btn.style.color = 'rgba(255, 255, 255, 0.8)';
            });
        }
    });

    // Fix file uploader buttons
    const uploaderBtns = doc.querySelectorAll('[data-testid="stFileUploader"] button');
    uploaderBtns.forEach(btn => {
        const txt = btn.textContent.trim().toLowerCase();
        // Hide the garbled "adc"/"add" more button
        if (txt.includes('adc') || txt.includes('add') || txt === '+') {
            btn.style.setProperty('font-size', '0', 'important');
            btn.style.setProperty('width', '36px', 'important');
            btn.style.setProperty('min-width', '36px', 'important');
            btn.style.setProperty('height', '36px', 'important');
            btn.style.setProperty('padding', '0', 'important');
            btn.style.setProperty('background', 'rgba(255,255,255,0.08)', 'important');
            btn.style.setProperty('border', '1px solid rgba(255,255,255,0.15)', 'important');
            btn.style.setProperty('border-radius', '8px', 'important');
            btn.style.setProperty('display', 'flex', 'important');
            btn.style.setProperty('align-items', 'center', 'important');
            btn.style.setProperty('justify-content', 'center', 'important');
            btn.innerHTML = '<span style="font-size:1.2rem;color:rgba(255,255,255,0.6)">+</span>';
        }
        // Fix the main upload/browse button
        if (txt.includes('upload') || txt.includes('browse')) {
            btn.style.setProperty('background', 'rgba(255,255,255,0.1)', 'important');
            btn.style.setProperty('border', '1px solid rgba(255,255,255,0.15)', 'important');
            btn.style.setProperty('color', 'rgba(255,255,255,0.8)', 'important');
            btn.style.setProperty('border-radius', '8px', 'important');
        }
    });
})();
</script>
""", height=0)
